# 카탈로그 전사 JSON(batch*.json) -> data/catalog-2026-n3.json + data/wines.xlsx + 보틀 이미지 복사
# 일회성 부트스트랩. 이후 데이터 수정은 data/wines.xlsx 에서 하고 build_data.py 를 실행하면 된다.
# 실행: python scripts/bootstrap_catalog.py <batch_json_dir> <crops_dir>
import json, glob, os, re, shutil, sys, unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from wine_utils import variety_tags

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
batch_dir, crops_dir = sys.argv[1], sys.argv[2]

producers, wines = [], []
for f in sorted(glob.glob(os.path.join(batch_dir, "batch*.json"))):
    d = json.load(open(f, encoding="utf-8"))
    producers += d["producers"]; wines += d["wines"]
wines.sort(key=lambda w: (w["page"], w["slot"]))
pids = {p["id"] for p in producers}
missing = {w["producer"] for w in wines} - pids
assert not missing, f"생산자 id 누락: {missing}"


def slugify(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s


def map_type(t):
    if "논알코올" in t: return "nonalcoholic"
    if "주정강화" in t: return "fortified"
    if "스파클링" in t: return "sparkling"
    if "로제" in t: return "rose"
    if "스위트" in t: return "sweet"
    if "레드" in t: return "red"
    return "white"


def style_profile(w, t):
    """종류·품종·도수·산지 키워드로 4축(바디/당도/산도/타닌) 초기값 산출. 엑셀에서 수정 가능."""
    name = w["name"].lower(); ko = w["nameKo"]; var = w["varieties"]; region = w["region"]; abv = w["abv"] or 0
    typestr = w["type"]
    if t == "red": b, s, a, tn = 3, 1, 3, 3
    elif t == "white": b, s, a, tn = 2, 1, 4, 1
    elif t == "rose": b, s, a, tn = 2, 2, 4, 1
    elif t == "sparkling": b, s, a, tn = 2, 1, 5, 1
    elif t == "fortified": b, s, a, tn = 4, 5, 2, 1
    elif t == "sweet": b, s, a, tn = 3, 5, 4, 1
    else: b, s, a, tn = (2, 2, 3, 2) if "레드" in typestr else (1, 2, 3, 1)  # 논알코올
    if t == "sparkling" and ("스위트" in typestr or "dolce" in name or "moscato" in name): s, a, b = 5, 3, 2
    if "세미스위트" in typestr: s = 3
    if "extra dry" in name: s += 1
    if any(k in var for k in ["카베르네 소비뇽", "타나", "템프라니요", "몬테풀치아노", "말벡", "바가", "쁘띠 베르도", "카르미네르", "나시오날", "나씨오날", "나쇼날"]): tn += 1
    if any(k in var for k in ["피노누아", "피노 누아"]) and t == "red": tn -= 1; b -= 1
    if any(k in var for k in ["프리미티보", "네그로아마로", "가르나차", "알리칸테 부쉐"]) or "아파시멘토" in ko: b += 1
    if "아파시멘토" in ko: s += 1
    if any(k in name for k in ["reserva", "riserva", "reserve", "grand", "gran ", "superyor", "mayor"]) and t == "red": b += 1
    if any(k in var for k in ["리슬링", "알바리뇨", "소비뇽 블랑", "베르멘티노", "아린토", "루레이로", "베르디키오", "산지오베제", "아잘", "엔크루자도"]): a += 1
    if any(k in region for k in ["비뉴 베르데", "비뉴베르데"]): a += 1; b -= 1
    if any(k in region for k in ["샹파뉴", "프란치아코르타"]): a = 5
    if any(k in var for k in ["비오니에", "샤도네이"]) and t == "white": b += 1
    if t in ("red", "white") and abv >= 14.5: b += 1
    if t in ("red", "white") and 0 < abv <= 11: b -= 1
    if "모스카텔" in var and t == "white": s += 1
    clamp = lambda v: max(1, min(5, v))
    return clamp(b), clamp(s), clamp(a), clamp(tn)


rows, used = [], set()
for i, w in enumerate(wines, 1):
    slug = slugify(w["name"]) or f"wine-{i}"
    if slug in used: slug = f"{slug}-{w['vintage'].lower()}"
    assert slug not in used, slug
    used.add(slug)
    t = map_type(w["type"])
    b, s, a, tn = style_profile(w, t)
    rows.append({
        "id": slug, "order": i, "catalogPage": w["page"], "producer": w["producer"],
        "nameKo": w["nameKo"], "name": w["name"], "vintage": w["vintage"], "type": t, "typeLabel": w["type"],
        "country": w["country"], "region": w["region"], "varieties": w["varieties"], "volume": w["volume"], "abv": w["abv"],
        "vivino": w["vivino"], "scores": w["scores"], "awards": w["awards"], "badges": w["badges"], "note": w["note"],
        "body": b, "sweetness": s, "acidity": a, "tannin": tn, "vivinoUrl": "",
        "profileStatus": "estimated", "varietyTags": ", ".join(variety_tags(w["varieties"])),
    })
    src = os.path.join(crops_dir, f"p{w['page']:02d}-s{w['slot']}.png")
    dst = os.path.join(ROOT, "public", "images", "wines", f"{slug}.png")
    shutil.copyfile(src, dst)

os.makedirs(os.path.join(ROOT, "data"), exist_ok=True)
json.dump({"source": "올빈와인 와인리스트 2026 N°3 (공급가 제외 전사)", "producers": producers, "wines": rows},
          open(os.path.join(ROOT, "data", "catalog-2026-n3.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)

# ---- 마스터 엑셀 ----
WINE_COLS = [
    ("id", "id (파일명, 수정 금지)"), ("order", "정렬"), ("catalogPage", "카탈로그 쪽"), ("producer", "생산자 id"),
    ("nameKo", "와인명(한글)"), ("name", "와인명(원어)"), ("vintage", "빈티지"), ("type", "타입코드"), ("typeLabel", "종류(카탈로그 표기)"),
    ("country", "국가"), ("region", "지역"), ("varieties", "품종"), ("volume", "용량"), ("abv", "도수(%)"),
    ("vivino", "Vivino 점수"), ("scores", "평론가 점수"), ("awards", "수상 (줄바꿈 구분)"), ("badges", "뱃지 (쉼표 구분)"), ("note", "한 줄 소개"),
    ("body", "바디 1-5"), ("sweetness", "당도 1-5"), ("acidity", "산도 1-5"), ("tannin", "타닌 1-5"), ("vivinoUrl", "Vivino URL (비우면 자동 검색링크)"),
    ("profileStatus", "프로필 상태 (estimated=추정 / verified=시음 검수)"), ("varietyTags", "품종 태그 (쉼표 구분, 비우면 품종 컬럼에서 자동 생성)"),
]
PROD_COLS = [("id", "id (수정 금지)"), ("name", "생산자(원어)"), ("nameKo", "생산자(한글)"), ("country", "국가"), ("region", "지역"), ("story", "소개")]

wb = Workbook()
ws = wb.active; ws.title = "wines"
head_font = Font(bold=True); head_fill = PatternFill("solid", start_color="F3EAE4", end_color="F3EAE4")
ws.append([h for _, h in WINE_COLS])
for c in ws[1]: c.font = head_font; c.fill = head_fill
for r in rows:
    ws.append([("\n".join(r[k]) if k == "awards" else ", ".join(r[k]) if k == "badges" else r[k]) for k, _ in WINE_COLS])
widths = {"A": 34, "E": 34, "F": 38, "I": 20, "K": 22, "L": 40, "P": 22, "Q": 46, "S": 40, "X": 30}
for col, wdt in widths.items(): ws.column_dimensions[col].width = wdt
for row in ws.iter_rows(min_row=2):
    for c in row: c.alignment = Alignment(wrap_text=True, vertical="top")
ws.freeze_panes = "E2"

ws2 = wb.create_sheet("producers")
ws2.append([h for _, h in PROD_COLS])
for c in ws2[1]: c.font = head_font; c.fill = head_fill
for p in producers: ws2.append([p[k] for k, _ in PROD_COLS])
for col, wdt in {"A": 24, "B": 30, "C": 22, "D": 10, "E": 22, "F": 100}.items(): ws2.column_dimensions[col].width = wdt
for row in ws2.iter_rows(min_row=2):
    for c in row: c.alignment = Alignment(wrap_text=True, vertical="top")

ws3 = wb.create_sheet("안내")
for line in [
    "올빈와인 앱 마스터 데이터",
    "- wines 시트: 와인 1행 = 앱의 와인 1개. 공급가/소비자가 컬럼은 의도적으로 없음 (앱에 가격 미노출).",
    "- 바디/당도/산도/타닌 1-5: 취향 찾기 퀴즈의 근거. 초기값은 종류·품종·도수 규칙으로 자동 산출한 것이므로 시음 판단으로 조정할 것.",
    "- 타입코드: red / white / rose / sparkling / fortified / sweet / nonalcoholic",
    "- 뱃지: BEST, 추천, NEW, BIO, EXCLUSIVE, LIMITED, SOLD OUT, COMING SOON 등 쉼표로 구분.",
    "- 프로필 상태: estimated(규칙으로 산출한 추정치, 앱에 '검수 전' 표시) / verified(시음 검수 완료). 검수한 와인만 verified로 바꿀 것.",
    "- 품종 태그: 비워두면 품종 컬럼에서 자동 생성됨. 필터에 쓰이므로 표기를 통일할 것 (예: 피노 누아, 토우리가 나시오날).",
    "- 보틀 이미지: public/images/wines/<id>.png. 실제 제품 사진으로 교체할 때 같은 파일명으로 덮어쓰면 됨.",
    "- 수정 후: python scripts/build_data.py 실행 -> src/data/*.json 갱신 -> git push 하면 자동 배포.",
]: ws3.append([line])
ws3.column_dimensions["A"].width = 110
wb.save(os.path.join(ROOT, "data", "wines.xlsx"))
print(f"wines {len(rows)}, producers {len(producers)} -> data/wines.xlsx, data/catalog-2026-n3.json, public/images/wines/*.png")
