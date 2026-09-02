# data/wines.xlsx (마스터) -> src/data/wines.json, src/data/producers.json
# 읽기 전용으로만 연다 (openpyxl로 기존 xlsx를 다시 저장하지 않음).
# 실행: python scripts/build_data.py
import json, os, re, sys, urllib.parse
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from wine_utils import variety_tags

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "data", "wines.xlsx")
OUT = os.path.join(ROOT, "src", "data")
TYPES = {"red", "white", "rose", "sparkling", "fortified", "sweet", "nonalcoholic"}
STATUSES = {"estimated", "verified"}
BADGES = {"BEST", "추천", "NEW", "BIO", "EXCLUSIVE", "LIMITED", "SOLD OUT", "COMING SOON"}
ID_RE = re.compile(r"[a-z0-9-]+")

# 이 repo는 공개다 — 가격·재고류 컬럼이 마스터 엑셀에 유입되면 빌드를 즉시 중단한다.
PRICE_RE = re.compile(r"공급|소비자가|단가|원가|판매가|장터가|할인가|마진|재고|price|cost|margin|stock", re.I)

# 엑셀 1행 헤더 문구 -> 내부 키. 컬럼 "위치"가 아니라 "이름"으로 읽으므로
# 컬럼 순서를 바꾸거나 사이에 끼워 넣어도 안전하다. 헤더 문구를 바꾸면 여기도 같이 수정할 것.
WINE_HEADERS = {
    "id (파일명, 수정 금지)": "id",
    "정렬": "order",
    "카탈로그 쪽": "catalogPage",
    "생산자 id": "producer",
    "와인명(한글)": "nameKo",
    "와인명(원어)": "name",
    "빈티지": "vintage",
    "타입코드": "type",
    "종류(카탈로그 표기)": "typeLabel",
    "국가": "country",
    "지역": "region",
    "품종": "varieties",
    "용량": "volume",
    "도수(%)": "abv",
    "Vivino 점수": "vivino",
    "평론가 점수": "scores",
    "수상 (줄바꿈 구분)": "awards",
    "뱃지 (쉼표 구분)": "badges",
    "한 줄 소개": "note",
    "바디 1-5": "body",
    "당도 1-5": "sweetness",
    "산도 1-5": "acidity",
    "타닌 1-5": "tannin",
    "Vivino URL (비우면 자동 검색링크)": "vivinoUrl",
    "프로필 상태 (estimated=추정 / verified=시음 검수)": "profileStatus",
    "품종 태그 (쉼표 구분, 비우면 품종 컬럼에서 자동 생성)": "varietyTags",
}
PROD_HEADERS = {
    "id (수정 금지)": "id",
    "생산자(원어)": "name",
    "생산자(한글)": "nameKo",
    "국가": "country",
    "지역": "region",
    "소개": "story",
}
PROD_KEYS = ["id", "name", "nameKo", "country", "region", "story"]

wb = load_workbook(XLSX, read_only=True, data_only=True)
errors = []


def read_sheet(name, header_map):
    ws = wb[name]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    cols = {}
    for j, h in enumerate(header):
        if not h:
            continue
        if PRICE_RE.search(h):
            errors.append(f"{name} 시트: 가격·재고류 컬럼 '{h}' 발견 — 이 repo는 공개이므로 가격 정보 금지. 컬럼을 제거할 것")
            continue
        if h not in header_map:
            errors.append(f"{name} 시트: 알 수 없는 컬럼 '{h}' — 헤더 문구를 바꿨다면 build_data.py 의 매핑도 같이 수정할 것")
            continue
        key = header_map[h]
        if key in cols:
            errors.append(f"{name} 시트: 컬럼 '{h}' 중복")
        cols[key] = j
    missing = [h for h, k in header_map.items() if k not in cols]
    if missing:
        errors.append(f"{name} 시트: 필수 컬럼 없음 — {', '.join(missing)}")
    if errors:  # 헤더가 어긋난 상태로 행을 읽으면 엉뚱한 오류가 쏟아지므로 여기서 멈춤
        return []
    out = []
    for i, row in enumerate(rows, start=2):
        vals = {k: (row[j] if j < len(row) else None) for k, j in cols.items()}
        if all(v is None or str(v).strip() == "" for v in vals.values()):
            continue  # 완전 빈 행만 조용히 건너뜀
        if vals["id"] is None or str(vals["id"]).strip() == "":
            errors.append(f"{name} {i}행: id 비어 있음 — 데이터가 남아 있는 행은 id 필수 (한 종이 통째로 누락되는 사고 방지)")
            continue
        vals["_row"] = i
        out.append(vals)
    return out


producers = read_sheet("producers", PROD_HEADERS)
wines = read_sheet("wines", WINE_HEADERS)
if errors:
    print("데이터 오류 — 수정 후 다시 실행:"); [print("  -", e) for e in errors]; sys.exit(1)

seen_producers = set()
for p in producers:
    r = p["_row"]
    if p["id"] in seen_producers:
        errors.append(f"producers {r}행: id 중복 '{p['id']}'")
    seen_producers.add(p["id"])

pids = {p["id"] for p in producers}
seen = set()
seen_orders = {}
for w in wines:
    r = w["_row"]
    if not ID_RE.fullmatch(str(w["id"])):
        errors.append(f"wines {r}행: id '{w['id']}' 는 소문자·숫자·하이픈만 허용")
    if w["id"] in seen: errors.append(f"wines {r}행: id 중복 '{w['id']}'")
    seen.add(w["id"])
    if not isinstance(w["order"], int) or isinstance(w["order"], bool):
        errors.append(f"wines {r}행: 정렬 값 '{w['order']}' 는 정수여야 함")
    elif w["order"] in seen_orders:
        errors.append(f"wines {r}행: 정렬 값 {w['order']} 이 {seen_orders[w['order']]}행과 중복")
    else:
        seen_orders[w["order"]] = r
    if w["producer"] not in pids: errors.append(f"wines {r}행: 생산자 id '{w['producer']}' 가 producers 시트에 없음")
    if w["type"] not in TYPES: errors.append(f"wines {r}행: 타입코드 '{w['type']}' 는 {sorted(TYPES)} 중 하나여야 함")
    for k in ("body", "sweetness", "acidity", "tannin"):
        v = w[k]
        is_integer = isinstance(v, (int, float)) and not isinstance(v, bool) and float(v).is_integer()
        if not is_integer or not 1 <= v <= 5:
            errors.append(f"wines {r}행: {k} 값 '{v}' 는 1~5 정수여야 함")
    for k in ("nameKo", "name", "vintage", "country", "region", "varieties"):
        if not w[k]: errors.append(f"wines {r}행: '{k}' 비어 있음")
    if w["abv"] not in (None, ""):
        if not isinstance(w["abv"], (int, float)) or not 0 <= float(w["abv"]) <= 25:
            errors.append(f"wines {r}행: 도수 '{w['abv']}' 는 0~25 사이 숫자여야 함")
    if w["vivino"] not in (None, ""):
        if not isinstance(w["vivino"], (int, float)) or not 0 <= float(w["vivino"]) <= 5:
            errors.append(f"wines {r}행: Vivino 점수 '{w['vivino']}' 는 0~5 사이 숫자여야 함")
    if w["vivinoUrl"] not in (None, "") and not str(w["vivinoUrl"]).startswith("http"):
        errors.append(f"wines {r}행: Vivino URL '{w['vivinoUrl']}' 은 http(s)로 시작해야 함")
    for b in [s.strip() for s in str(w["badges"] or "").split(",") if s.strip()]:
        if b not in BADGES:
            errors.append(f"wines {r}행: 뱃지 '{b}' 는 {sorted(BADGES)} 중 하나여야 함 (오타 확인)")
    if w["profileStatus"] not in (None, "") and w["profileStatus"] not in STATUSES:
        errors.append(f"wines {r}행: 프로필 상태 '{w['profileStatus']}' 는 estimated 또는 verified 여야 함")
    img = os.path.join(ROOT, "public", "images", "wines", f"{w['id']}.png")
    if not os.path.exists(img): errors.append(f"wines {r}행: 이미지 없음 public/images/wines/{w['id']}.png")
    webp = os.path.join(ROOT, "public", "images", "wines", f"{w['id']}.webp")
    if os.path.exists(img) and not os.path.exists(webp):
        errors.append(f"wines {r}행: WebP 없음 — 'npm run images' 를 실행해 표시용 WebP를 생성할 것")
if errors:
    print("데이터 오류 — 수정 후 다시 실행:"); [print("  -", e) for e in errors]; sys.exit(1)


def split_lines(v): return [s.strip() for s in str(v).splitlines() if s.strip()] if v else []
def split_comma(v): return [s.strip() for s in str(v).split(",") if s.strip()] if v else []
def vivino_search(w):
    q = w["name"] + (f" {w['vintage']}" if str(w["vintage"]).isdigit() else "")
    return "https://www.vivino.com/search/wines?q=" + urllib.parse.quote(q)


out_w = []
for w in sorted(wines, key=lambda x: (x["order"] or 0)):
    out_w.append({
        "id": w["id"], "order": int(w["order"] or 0), "catalogPage": int(w["catalogPage"] or 0), "producer": w["producer"],
        "nameKo": w["nameKo"], "name": w["name"], "vintage": str(w["vintage"]), "type": w["type"], "typeLabel": w["typeLabel"] or "",
        "country": w["country"], "region": w["region"], "varieties": w["varieties"], "volume": w["volume"] or "750ml",
        "abv": float(w["abv"]) if w["abv"] not in (None, "") else None,
        "vivino": float(w["vivino"]) if w["vivino"] not in (None, "") else None,
        "scores": w["scores"] or "", "awards": split_lines(w["awards"]), "badges": split_comma(w["badges"]), "note": w["note"] or "",
        "body": int(w["body"]), "sweetness": int(w["sweetness"]), "acidity": int(w["acidity"]), "tannin": int(w["tannin"]),
        "image": f"/images/wines/{w['id']}.png", "vivinoUrl": w["vivinoUrl"] or vivino_search(w),
        "profileStatus": w["profileStatus"] or "estimated",
        "varietyTags": split_comma(w["varietyTags"]) or variety_tags(w["varieties"]),
    })
out_p = [{k: (p[k] or "") for k in PROD_KEYS} for p in producers]
for p in out_p: p["wineCount"] = sum(1 for w in out_w if w["producer"] == p["id"])

os.makedirs(OUT, exist_ok=True)
json.dump(out_w, open(os.path.join(OUT, "wines.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)
json.dump(out_p, open(os.path.join(OUT, "producers.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)
wb.close()
print(f"OK: wines {len(out_w)}, producers {len(out_p)} -> src/data/wines.json, producers.json")
